import time
import statistics
import BaseXClient
import openpyxl
from openpyxl import Workbook
import os

# Dati connessione
HOST = "localhost"
PORT = 1984
USERNAME = "admin"
PASSWORD = "1234"
DATABASE = "dataset_25"

# Percorso per salvare i risultati
RESULTS_DIR = r"C:\Users\fabyp\OneDrive\Documents\Progetto DB2\basex\risultati"
RESULTS_FILE = os.path.join(RESULTS_DIR, "risultati_query.xlsx")

# Dizionario delle query ottimizzate
query_list = {
    "Q1": """
  for $p in //Persona[Professione/NomeProfessione = "Medical sales representative"]
let $fonte := //Fonte[@ID_Fonte = $p/@ID_Fonte][xs:double(Affidabilita) > 0.5]
let $doc := //Documento[@ID_Documento = $p/@ID_Documento][xs:date(Scadenza) > xs:date("2028-01-01")]
where exists($fonte) and exists($doc)
return <Result>
  <Nome>{$p/Nome/text()}</Nome>
  <Cognome>{$p/Cognome/text()}</Cognome>
  <NomeFonte>{$fonte/NomeFonte/text()}</NomeFonte>
  <NazioneFonte>{$fonte/Nazione/text()}</NazioneFonte>
  <ScadenzaDocumento>{$doc/Scadenza/text()}</ScadenzaDocumento>
  <CittaDocumento>{$doc/Indirizzo/Citta/text()}</CittaDocumento>
</Result>



    """,

    "Q2": """
     for $p in //Persona[Professione/NomeProfessione = "Medical sales representative"]
let $fonte := //Fonte[@ID_Fonte = $p/@ID_Fonte][xs:double(Affidabilita) > 0.5]
let $doc := //Documento[@ID_Documento = $p/@ID_Documento][xs:date(Scadenza) > xs:date("2028-01-01")]
where exists($fonte) and exists($doc)
return <Result>
  <Nome>{$p/Nome/text()}</Nome>
  <Cognome>{$p/Cognome/text()}</Cognome>
  <NomeFonte>{$fonte/NomeFonte/text()}</NomeFonte>
  <NazioneFonte>{$fonte/Nazione/text()}</NazioneFonte>
  <ScadenzaDocumento>{$doc/Scadenza/text()}</ScadenzaDocumento>
  <CittaDocumento>{$doc/Indirizzo/Citta/text()}</CittaDocumento>
</Result>



    """,

    "Q3": """
       for $p in //Persona
where $p/@ID_Locale = $p/@ID_Documento or $p/@ID_Locale = $p/@ID_Fonte
return 
  <PersonaIDIdentici>
    {$p/@ID_Locale}
    <Nome>{$p/Nome/text()}</Nome>
    <Cognome>{$p/Cognome/text()}</Cognome>
    <ID_Documento>{$p/@ID_Documento}</ID_Documento>
    <ID_Fonte>{$p/@ID_Fonte}</ID_Fonte>
  </PersonaIDIdentici>


    """,

    "Q4": """
       for $d1 in //Documento
for $d2 in //Documento
where $d1/@ID_Documento != $d2/@ID_Documento
  and $d1/@ID_Fonte != $d2/@ID_Fonte
  and $d1/Indirizzo/Via = $d2/Indirizzo/Via
  and $d1/Indirizzo/Citta = $d2/Indirizzo/Citta
  and $d1/Indirizzo/CAP = $d2/Indirizzo/CAP
return
  <IndirizzoDuplicato>
    <Documento1>{$d1/@ID_Documento}</Documento1>
    <Documento2>{$d2/@ID_Documento}</Documento2>
    <Via>{$d1/Indirizzo/Via/text()}</Via>
    <Citta>{$d1/Indirizzo/Citta/text()}</Citta>
    <CAP>{$d1/Indirizzo/CAP/text()}</CAP>
  </IndirizzoDuplicato>



    """
}

def initialize_excel_file():
    """Crea o inizializza il file Excel con le intestazioni"""
    if not os.path.exists(RESULTS_DIR):
        os.makedirs(RESULTS_DIR)
    
    if os.path.exists(RESULTS_FILE):
        wb = openpyxl.load_workbook(RESULTS_FILE)
    else:
        wb = Workbook()
        # Rimuovi il foglio predefinito se vuoto
        if len(wb.sheetnames) == 1 and wb.active.max_row == 1 and wb.active.max_column == 1:
            wb.remove(wb.active)
    
    # Crea fogli per i risultati se non esistono
    for query_name in query_list.keys():
        # Foglio per la prima esecuzione
        if f"{query_name}_first" not in wb.sheetnames:
            ws = wb.create_sheet(title=f"{query_name}_first")
            ws.append(["Timestamp", "Tempo (ms)"])
        
        # Foglio per la media delle 30 esecuzioni
        if f"{query_name}_avg" not in wb.sheetnames:
            ws = wb.create_sheet(title=f"{query_name}_avg")
            ws.append(["Timestamp", "Tempo medio (ms)", "Deviazione standard"])
    
    wb.save(RESULTS_FILE)
    return wb

def save_results_to_excel(query_name, first_exec_time, avg_exec_time, std_dev):
    """Salva i risultati nel file Excel"""
    wb = openpyxl.load_workbook(RESULTS_FILE)
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    
    # Salva i risultati della prima esecuzione
    ws_first = wb[f"{query_name}_first"]
    ws_first.append([timestamp, first_exec_time])
    
    # Salva i risultati delle esecuzioni successive
    ws_avg = wb[f"{query_name}_avg"]
    ws_avg.append([timestamp, avg_exec_time, std_dev])
    
    wb.save(RESULTS_FILE)

def run_query_once(xquery):
    session = BaseXClient.Session(HOST, PORT, USERNAME, PASSWORD)
    try:
        session.execute(f"open {DATABASE}")
        start = time.perf_counter()
        result = session.execute(f"xquery {xquery}")
        end = time.perf_counter()
        elapsed = (end - start) * 1000  # ms
        return elapsed, result
    finally:
        session.close()

def run_query_30_times(xquery):
    times = []
    for i in range(30):
        session = BaseXClient.Session(HOST, PORT, USERNAME, PASSWORD)
        try:
            session.execute(f"open {DATABASE}")
            start = time.perf_counter()
            session.execute(f"xquery {xquery}")
            end = time.perf_counter()
            elapsed = (end - start) * 1000
            times.append(elapsed)
            print(f"    Esecuzione {i + 1:02d}: {elapsed:.2f} ms")
        finally:
            session.close()
    return times

def benchmark_query(name, xquery):
    print(f"\nBenchmark per {name}:")

    # Esecuzione singola
    tempo_singolo, result = run_query_once(xquery)
    print(f"  Esecuzione singola: {tempo_singolo:.2f} ms")
    if result.strip():
        print(f"  Risultato: {result.strip()[:100]}{'...' if len(result.strip()) > 100 else ''}")
    else:
        print("  Risultato: (vuoto)")

    # Esecuzioni ripetute
    print("  Esecuzioni ripetute (con connessione chiusa tra ogni esecuzione):")
    tempi = run_query_30_times(xquery)
    media = sum(tempi) / len(tempi)
    dev_std = statistics.stdev(tempi) if len(tempi) > 1 else 0.0
    print(f"  Media: {media:.2f} ms | Deviazione standard: {dev_std:.2f} ms")

    # Salva i risultati
    save_results_to_excel(name, tempo_singolo, media, dev_std)

def main():
    initialize_excel_file()
    for nome, query in query_list.items():
        benchmark_query(nome, query)

if __name__ == "__main__":
    main()